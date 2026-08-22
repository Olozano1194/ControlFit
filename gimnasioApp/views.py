from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated, AllowAny
from django.contrib.auth import authenticate
from django.contrib.auth import get_user_model
from rest_framework_simplejwt.views import TokenObtainPairView, TokenRefreshView
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework_simplejwt.exceptions import TokenError, InvalidToken
from .auth_cookie import set_refresh_cookie, clear_refresh_cookie
from .serializers import UsuarioSerializer, UsuarioGymSerializer, UsuarioGymDaySerializer, MembresiasSerializer, MembresiaAsignadaSerializer, PagoMembresiaSerializer, TipoEventoSerializer, EventoCalendarioSerializer, NotificationSerializer
from .models import Usuario, UsuarioGym, UsuarioGymDay, Membresia, MembresiaAsignada, PagoMembresia, Gimnasio, TipoEvento, EventoCalendario, Notification
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404
#para la imagen
from rest_framework.parsers import MultiPartParser, FormParser
#Para las filtraciones en la base de datos
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter
#Para las notificaciones
from datetime import datetime, date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
# Importar permisos personalizados
from .permissions import IsAdminUser, IsRecepcionUser, IsSuperAdmin
from .mixins import MultiTenantViewSetMixin
from .services.notifications import NotificationManager
from decimal import Decimal
from django.db.models import Sum
from rest_framework.decorators import action
from rest_framework.pagination import PageNumberPagination


# ============================================================
# PLATFORM PAGINATION
# ============================================================

class PlatformPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 100
    page_query_param = 'page'


# ============================================================
# VIEWSETS
# ============================================================

class UserViewSet(MultiTenantViewSetMixin, viewsets.ModelViewSet):
    queryset = Usuario.objects.all()
    serializer_class = UsuarioSerializer
    permission_classes = [IsAuthenticated, IsAdminUser]  # Solo admins pueden gestionar usuarios
    parser_classes = (MultiPartParser, FormParser)

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if not serializer.is_valid():
            return Response({
                "error": "Datos inválidos",
                "details": serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
        self.perform_create(serializer)
        user_data = serializer.instance

        return Response({
            "user": UsuarioSerializer(user_data, context={'request': request}).data,
        }, status=status.HTTP_201_CREATED)
    
    def update(self, request, *args, **kwargs):
        try:
            partial = kwargs.pop('partial', False)
            instance = self.get_object()



            serializer = self.get_serializer(
                instance, 
                data=request.data, 
                partial=partial,
                context={'request': request}
            )

            if serializer.is_valid():
                self.perform_update(serializer)
                return Response(serializer.data, status=status.HTTP_200_OK)
            
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            return Response({"error": "Error inesperado"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# Registro público para crear usuario inicial
# El login se maneja con SimpleJWT en /token/ (TokenObtainPairView)
class RegisterViewSet(APIView):
    permission_classes = [AllowAny]
    
    def post(self, request, *args, **kwargs):
        email = request.data.get('email')
        password = request.data.get('password')
        name = request.data.get('name')
        lastname = request.data.get('lastname')
        
        if not email or not password or not name or not lastname:
            return Response({
                'error': 'Todos los campos son requeridos',
                'required': ['email', 'password', 'name', 'lastname']
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Verificar si el usuario ya existe
        if get_user_model().objects.filter(email=email).exists():
            return Response({'error': 'El correo ya está registrado'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Crear gimnasio automáticamente
        email_prefix = email.split('@')[0].replace('.', ' ').title()
        gimnasio = Gimnasio.objects.create(name=f"Gimnasio {email_prefix}")
        
        # Crear usuario con gimnasio
        user = get_user_model()(
            email=email,
            name=name,
            lastname=lastname,
            roles='admin',  # Primer usuario es admin
            gimnasio=gimnasio
        )
        user.set_password(password)
        user.save()
        
        # Generar tokens JWT
        refresh = RefreshToken.for_user(user)
        
        response = Response({
            'message': 'Usuario creado exitosamente',
            'user': UsuarioSerializer(user, context={'request': request}).data,
            'access': str(refresh.access_token),
        }, status=status.HTTP_201_CREATED)
        
        # Establecer refresh token como cookie HttpOnly (helper compartido)
        set_refresh_cookie(response, refresh)
        
        return response


# ============================================================
# AUTH JWT — LOGIN / REFRESH / LOGOUT (cookie-based)
# ============================================================

class CookieTokenObtainPairView(TokenObtainPairView):
    """Login: valida credenciales, setea el refresh como cookie HttpOnly y devuelve solo access."""
    permission_classes = [AllowAny]

    def post(self, request, *args, **kwargs):
        response = super().post(request, *args, **kwargs)
        if response.status_code == status.HTTP_200_OK:
            refresh = response.data.get('refresh')
            if refresh:
                set_refresh_cookie(response, refresh)
                del response.data['refresh']  # Nunca exponer el refresh en el body
        return response


class CookieTokenRefreshView(TokenRefreshView):
    """Refresh: lee el refresh de la COOKIE (no del body), rota y re-setea la cookie."""
    permission_classes = [AllowAny]

    def post(self, request, *args, **kwargs):
        refresh_token = request.COOKIES.get('refresh_token')
        if not refresh_token:
            return Response({'detail': 'No refresh token'}, status=status.HTTP_401_UNAUTHORIZED)

        # Inyectar el refresh desde la cookie directo al serializer (mismo flujo que TokenViewBase)
        serializer = self.get_serializer(data={'refresh': refresh_token})
        try:
            serializer.is_valid(raise_exception=True)
        except TokenError as e:
            raise InvalidToken(e.args[0])

        response = Response(serializer.validated_data, status=status.HTTP_200_OK)

        new_refresh = response.data.get('refresh')
        if new_refresh:
            set_refresh_cookie(response, new_refresh)
            del response.data['refresh']
        return response


class LogoutView(APIView):
    """Logout: blacklistea el refresh de la cookie y limpia la cookie (idempotente)."""
    permission_classes = [AllowAny]

    def post(self, request):
        refresh_token = request.COOKIES.get('refresh_token')
        if refresh_token:
            try:
                RefreshToken(refresh_token).blacklist()
            except Exception:
                # Token inválido/expirado/ya blacklistado: seguimos y limpiamos igual
                pass
        response = Response({'detail': 'Logged out'})
        clear_refresh_cookie(response)
        return response


class userProfileView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            user_serializer = UsuarioSerializer(request.user, context={'request': request})
            user_data = user_serializer.data            
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        return Response({'user': user_data}, status=status.HTTP_200_OK)
          
    def list(self, request):
        try:
            gimnasio = request.gimnasio
            if gimnasio:
                users = Usuario.objects.filter(gimnasio=gimnasio).order_by('-id')
            else:
                users = Usuario.objects.none()
            serializer = UsuarioSerializer(users, many=True, context={'request': request})
            return Response({'users': serializer.data}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)    


# ============================================================
# MIEMBROS DEL GIMNASIO
# ============================================================

class UsuarioGymViewSet(MultiTenantViewSetMixin, viewsets.ModelViewSet):
    queryset = UsuarioGym.objects.all()
    serializer_class = UsuarioGymSerializer
    permission_classes = [IsAuthenticated, IsRecepcionUser]  # Admin y recepcion pueden acceder
    filter_backends = [DjangoFilterBackend, SearchFilter]
    search_fields = ['name', 'lastname']


class UsuarioGymDayViewSet(MultiTenantViewSetMixin, viewsets.ModelViewSet):
    queryset = UsuarioGymDay.objects.all()
    serializer_class = UsuarioGymDaySerializer
    permission_classes = [IsAuthenticated, IsRecepcionUser]  # Admin y recepcion pueden acceder
    filter_backends = [DjangoFilterBackend, SearchFilter]
    search_fields = ['name', 'lastname']

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if not serializer.is_valid():
            return Response({
                "error": "Datos inválidos para ingreso diario",
                "details": serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
        self.perform_create(serializer)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


# ============================================================
# ESTADÍSTICAS DEL DASHBOARD (Miembros Activos, Nuevos, Retention)
# ============================================================

class DashboardStatsView(APIView):
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        from django.db.models import Count, Q
        from datetime import date, timedelta
        
        gimnasio = request.gimnasio
        
        if not gimnasio:
            return Response({
                'active_members': 0,
                'new_today': 0,
                'retention_rate': 0
            })
        
        today = date.today()
        
        # 1. Miembros activos (membresías vigentes)
        active_memberships = MembresiaAsignada.objects.filter(
            miembro__gimnasio=gimnasio,
            dateInitial__lte=today,
            dateFinal__gte=today
        )
        active_count = active_memberships.count()
        
        # 2. Nuevos hoy (miembros que se registraronHOY)
        new_today = active_memberships.filter(
            dateInitial=today
        ).count()
        
        # 3. Retention Rate
        # Mes anterior
        if today.month == 1:
            prev_month = 12
            prev_year = today.year - 1
        else:
            prev_month = today.month - 1
            prev_year = today.year
        
        # Primer día del mes anterior
        prev_month_start = date(prev_year, prev_month, 1)
        
        # Último día del mes anterior
        if prev_month == 12:
            prev_month_end = date(prev_year + 1, 1, 1) - timedelta(days=1)
        else:
            prev_month_end = date(prev_year, prev_month + 1, 1) - timedelta(days=1)
        
        # Miembros únicos activos el mes anterior (distinct por miembro)
        prev_active_ids = MembresiaAsignada.objects.filter(
            miembro__gimnasio=gimnasio,
            dateInitial__lte=prev_month_end,
            dateFinal__gte=prev_month_start
        ).values_list('miembro', flat=True).distinct()
        prev_active_count = len(prev_active_ids)
        
        # De esos miembros, cuántos siguen activos hoy (con cualquier membresía)
        still_active_count = 0
        if prev_active_count > 0:
            still_active_count = MembresiaAsignada.objects.filter(
                miembro__in=list(prev_active_ids),
                miembro__gimnasio=gimnasio,
                dateFinal__gte=today,
                dateInitial__lte=today
            ).values('miembro').distinct().count()
        
        # Calcular retention
        if prev_active_count > 0:
            retention = round((still_active_count / prev_active_count) * 100, 1)
        else:
            retention = 100.0  # Si no hay miembros anteriores, 100%
        
        return Response({
            'active_members': active_count,
            'new_today': new_today,
            'retention_rate': retention
        })


# ============================================================
# HOME / DASHBOARD
# ============================================================

class Home(APIView):
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        gimnasio = request.gimnasio
        
        if gimnasio:
            UserGymList = MembresiaAsignada.objects.filter(miembro__gimnasio=gimnasio).order_by('-id')
            UserDayList = UsuarioGymDay.objects.filter(gimnasio=gimnasio).order_by('-id')
        else:
            UserGymList = MembresiaAsignada.objects.none()
            UserDayList = UsuarioGymDay.objects.none()

        now = timezone.now()
        month = now.month
        year = now.year

        # Mes anterior
        if month == 1:
            prev_month, prev_year = 12, year - 1
        else:
            prev_month, prev_year = month - 1, year

        num_miembros = UserGymList.count()

        miembrosDay_mes = UserDayList.filter(dateInitial__month=month, dateInitial__year=year)
        total_day_mes = sum(user.price for user in miembrosDay_mes)
        # Pagos recibidos este mes (dinero real, no esperado)
        pagos_mes = PagoMembresia.objects.filter(
            membresia_asignada__miembro__gimnasio=gimnasio,
            fecha_pago__month=month,
            fecha_pago__year=year
        ).aggregate(total=Sum('monto'))['total'] or Decimal('0')
        total_month = pagos_mes + total_day_mes
        miembros_mes_count = UserGymList.filter(dateInitial__month=month, dateInitial__year=year).count()

        # Miembros del mes anterior
        miembros_mes_anterior = UserGymList.filter(
            dateInitial__month=prev_month,
            dateInitial__year=prev_year
        ).count()

        # Diferencia vs mes anterior
        diff_miembros = miembros_mes_count - miembros_mes_anterior

        total_day = sum(user.price for user in UserDayList)
        # Dinero real recibido: suma de todos los pagos de membresias + ingresos diarios
        pagos_totales = PagoMembresia.objects.filter(
            membresia_asignada__miembro__gimnasio=gimnasio
        ).aggregate(total=Sum('monto'))['total'] or Decimal('0')
        total = pagos_totales + total_day

        # ---- Dashboard de cobranza ----
        today = date.today()
        active_memberships = UserGymList.filter(
            dateInitial__lte=today,
            dateFinal__gte=today
        )

        # Obtener total de pagos por membresia activa en una sola consulta
        pago_totals = PagoMembresia.objects.filter(
            membresia_asignada__in=active_memberships
        ).values('membresia_asignada').annotate(
            total_monto=Sum('monto')
        )
        pago_dict = {item['membresia_asignada']: item['total_monto'] for item in pago_totals}

        por_cobrar = Decimal('0')
        al_dia = 0
        con_deuda = 0

        for m in active_memberships:
            total_pagado = pago_dict.get(m.id, Decimal('0'))
            saldo = m.price - total_pagado
            if saldo > 0:
                por_cobrar += saldo
                con_deuda += 1
            else:
                al_dia += 1

        return JsonResponse({ 
            'num_miembros': num_miembros, 
            'total_month': float(total_month), 
            'miembros_mes': miembros_mes_count,
            'total': float(total),
            'miembros_mes_anterior': miembros_mes_anterior,   
            'diff_miembros': diff_miembros,
            'por_cobrar': float(por_cobrar),
            'al_dia': al_dia,
            'con_deuda': con_deuda,
        })
    

# ============================================================
# MEMBRESIAS
# ============================================================

class MembresiaViewSet(MultiTenantViewSetMixin, viewsets.ModelViewSet):
    queryset = Membresia.objects.all()
    serializer_class = MembresiasSerializer
    permission_classes = [IsAuthenticated, IsRecepcionUser]


class MembresiaAsignadaViewSet(MultiTenantViewSetMixin, viewsets.ModelViewSet):
    queryset = MembresiaAsignada.objects.all()
    serializer_class = MembresiaAsignadaSerializer
    permission_classes = [IsAuthenticated, IsRecepcionUser]
    filter_backends = [DjangoFilterBackend, SearchFilter]
    search_fields = ['miembro__name', 'miembro__lastname']
    gimnasio_field = 'miembro__gimnasio'

    def get_queryset(self):
        """Filtrar membresías asignadas por gimnasio del usuario actual."""
        queryset = super().get_queryset()
        
        # Filtro adicional por miembro si se pasa
        miembro_id = self.request.query_params.get('miembro')
        if miembro_id:
            queryset = queryset.filter(miembro_id=miembro_id)
        
        return queryset.order_by('-id')


# ============================================================
# PAGO MEMBRESIA (Nested under MemberShipsAsignada/{pk}/pagos/)
# ============================================================

class PagoMembresiaViewSet(viewsets.ModelViewSet):
    serializer_class = PagoMembresiaSerializer
    permission_classes = [IsAuthenticated, IsRecepcionUser]

    def get_membresia_asignada(self):
        pk = self.kwargs.get('pk')
        return get_object_or_404(
            MembresiaAsignada.objects.filter(miembro__gimnasio=self.request.gimnasio),
            pk=pk
        )

    def get_queryset(self):
        membresia = self.get_membresia_asignada()
        return PagoMembresia.objects.filter(
            membresia_asignada=membresia
        ).order_by('-fecha_pago')

    def perform_create(self, serializer):
        membresia = self.get_membresia_asignada()
        serializer.save(membresia_asignada=membresia)

    def get_serializer_context(self):
        context = super().get_serializer_context()
        try:
            membresia = self.get_membresia_asignada()
            context['membresia_asignada'] = membresia
        except Exception:
            pass
        return context


# ============================================================
# NOTIFICACIONES (Nivel 1) — API persistente multi-tenant
# ============================================================

class NotificationViewSet(MultiTenantViewSetMixin, viewsets.ReadOnlyModelViewSet):
    """API de notificaciones para admin y recepcionistas.

    El listado dispara la generación perezosa e idempotente de notificaciones
    (vencimientos de membresías y eventos del día) y devuelve SOLO las no
    leídas, ordenadas de más reciente a más antigua. Las leídas desaparecen.
    """
    queryset = Notification.objects.all()
    serializer_class = NotificationSerializer
    permission_classes = [IsAuthenticated, IsRecepcionUser]

    def get_queryset(self):
        """Solo notificaciones no leídas del gimnasio del request."""
        queryset = super().get_queryset().filter(is_read=False)
        return queryset.order_by('-created_at')

    def list(self, request, *args, **kwargs):
        # Generación perezosa e idempotente antes de responder
        NotificationManager.generate_for_gimnasio(request.gimnasio)
        return super().list(request, *args, **kwargs)

    @action(detail=True, methods=['post'], url_path='marcar-leida')
    def marcar_leida(self, request, pk=None):
        """Marca una notificación como leída (is_read=True, read_at=ahora)."""
        notification = self.get_object()
        notification.is_read = True
        notification.read_at = timezone.now()
        notification.save(update_fields=['is_read', 'read_at'])
        return Response({'status': 'ok'})

    @action(detail=False, methods=['post'], url_path='marcar-todas-leidas')
    def marcar_todas_leidas(self, request):
        """Marca todas las notificaciones no leídas del gimnasio como leídas."""
        now = timezone.now()
        # get_queryset ya limita al gimnasio del request y a las no leídas
        marked = self.get_queryset().update(is_read=True, read_at=now)
        return Response({'status': 'ok', 'marked': marked})

    @action(detail=False, methods=['get'], url_path='no-leidas')
    def no_leidas(self, request):
        """Conteo de notificaciones no leídas para el badge del frontend."""
        count = self.get_queryset().count()
        return Response({'count': count})

# ============================================================
# ACTIVIDADES RECIENTES
# ============================================================

class ActivitiesView(APIView):
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        gimnasio = request.gimnasio
        
        activities = []
        
        if gimnasio:
            membresias_recientes = MembresiaAsignada.objects.filter(
                miembro__gimnasio=gimnasio
            ).select_related('miembro', 'membresia').order_by('-created_at')[:5]

            ingresos = UsuarioGymDay.objects.filter(
                gimnasio=gimnasio
            ).order_by('-created_at')[:5]
            
            for membresia in membresias_recientes:
                
                activities.append({
                    'id': f'm-{membresia.id}',
                    'type': 'new_member',
                    'icon': 'person_add',
                    'color': 'primary',
                    'title': 'Nuevo miembro registrado',
                    'description': f'{membresia.miembro.name} {membresia.miembro.lastname} - {membresia.membresia.name}',                    
                    'created_at': membresia.created_at.isoformat(),
                    'time_ago': self.get_time_ago(membresia.created_at),
                })            
            
            
            for ingreso in ingresos:
                
                activities.append({
                    'id': f'i-{ingreso.id}',
                    'type': 'entry',
                    'icon': 'login',
                    'color': 'info',
                    'title': 'Ingreso registrado',
                    'description': f'{ingreso.name} {ingreso.lastname}',
                    'created_at': ingreso.created_at.isoformat(),
                    'amount': float(ingreso.price),
                    'time_ago': self.get_time_ago(ingreso.created_at)
                })

            # Pagos recientes (abonos o pagos completos)
            pagos_recientes = PagoMembresia.objects.filter(
                membresia_asignada__miembro__gimnasio=gimnasio
            ).select_related(
                'membresia_asignada__miembro',
                'membresia_asignada__membresia'
            ).order_by('-fecha_pago')[:5]

            for pago in pagos_recientes:
                asignacion = pago.membresia_asignada
                miembro = asignacion.miembro
                # Determinar si es pago completo o parcial
                if pago.monto >= asignacion.price:
                    tipo_pago = 'Pago completo'
                else:
                    tipo_pago = 'Abono'
                activities.append({
                    'id': f'p-{pago.id}',
                    'type': 'payment',
                    'icon': 'payments',
                    'color': 'success',
                    'title': f'{tipo_pago} - {pago.metodo_pago}',
                    'description': f'{miembro.name} {miembro.lastname} - {asignacion.membresia.name}',
                    'amount': float(pago.monto),
                    'created_at': pago.fecha_pago.isoformat(),
                    'time_ago': self.get_time_ago(pago.fecha_pago),
                })
        
        activities.sort(key=lambda x: x['created_at'], reverse=True)
        
        return Response(activities[:10])
    
     # ==========================================================
    # HELPERS
    # ==========================================================

    def ensure_datetime(self, date_obj):
        """Convierte date → datetime seguro"""
        if isinstance(date_obj, date) and not isinstance(date_obj, datetime):
            date_obj = datetime.combine(date_obj, datetime.min.time())

        if timezone.is_naive(date_obj):
            date_obj = timezone.make_aware(date_obj)

        return date_obj
    
    def get_time_ago(self, date_obj):
        now = timezone.now()
        diff = now - date_obj

        if diff.days > 0:
            return f'{diff.days} día{"s" if diff.days != 1 else ""}'
        elif diff.seconds >= 3600:
            hours = diff.seconds // 3600
            return f'{hours} hora{"s" if hours != 1 else ""}'
        elif diff.seconds >= 60:
            minutes = diff.seconds // 60
            return f'{minutes} minuto{"s" if minutes != 1 else ""}'
        else:
            return 'Ahora mismo'
        
# ============================================================
# EXPORTAR REPORTE
# ============================================================

class ExportReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        gimnasio = request.gimnasio

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"

        # ESTILOS
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        bold_font = Font(bold=True)
        center = Alignment(horizontal="center")

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # TÍTULO
        ws.merge_cells('A1:F1')
        ws['A1'] = 'REPORTE DEL GIMNASIO'
        ws['A1'].font = Font(size=18, bold=True)
        ws['A1'].alignment = center

        ws.append([])
        ws.append([f'Generado: {timezone.now().strftime("%d/%m/%Y %H:%M")}'])
        ws.append([])

        # DATOS
        if gimnasio:
            miembros = MembresiaAsignada.objects.filter(miembro__gimnasio=gimnasio)
            diarios = UsuarioGymDay.objects.filter(gimnasio=gimnasio)
        else:
            miembros = []
            diarios = []

        total_membresias = sum(m.price for m in miembros)
        total_diarios = sum(d.price for d in diarios)
        total = total_membresias + total_diarios

        # ==========================================================
        # ESTADÍSTICAS
        # ==========================================================
        ws.append(['ESTADÍSTICAS DEL MES'])

        headers = ['Total miembros', 'Ingresos membresías', 'Ingresos diarios', 'Total']
        ws.append(headers)

        for col_num, _ in enumerate(headers, 1):
            cell = ws.cell(row=ws.max_row, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        ws.append([
            len(miembros),
            total_membresias,
            total_diarios,
            total
        ])

        # formato dinero
        for col in range(2, 5):
            ws.cell(row=ws.max_row, column=col).number_format = '$#,##0'

        ws.append([])

        # ==========================================================
        # 👤 MIEMBROS
        # ==========================================================
        ws.append(['MIEMBROS'])
        headers = ['Nombre', 'Apellido', 'Membresía', 'Fecha Inicio', 'Precio']
        ws.append(headers)

        for col_num, _ in enumerate(headers, 1):
            cell = ws.cell(row=ws.max_row, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border

        for m in miembros[:50]:
            ws.append([
                m.miembro.name,
                m.miembro.lastname,
                m.membresia.name,
                m.dateInitial.strftime('%d/%m/%Y'),
                m.price
            ])

            row = ws.max_row
            ws.cell(row=row, column=5).number_format = '$#,##0'

        ws.append([])

        # ==========================================================
        # INGRESOS DIARIOS
        # ==========================================================
        ws.append(['INGRESOS DIARIOS'])
        headers = ['Nombre', 'Apellido', 'Fecha', 'Monto']
        ws.append(headers)

        for col_num, _ in enumerate(headers, 1):
            cell = ws.cell(row=ws.max_row, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border

        for d in diarios[:50]:
            ws.append([
                d.name,
                d.lastname,
                d.dateInitial.strftime('%d/%m/%Y'),
                d.price
            ])

            row = ws.max_row
            ws.cell(row=row, column=4).number_format = '$#,##0'

        # ==========================================================
        # AUTO AJUSTE COLUMNAS
        # ==========================================================
        from openpyxl.utils import get_column_letter
        
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)

            for cell in col:
                try:
                    # Ignorar celdas combinadas (MergedCell)
                    if cell.value and not isinstance(cell, type(ws.merged_cells)):
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            ws.column_dimensions[col_letter].width = max_length + 3 if max_length > 0 else 15

        # ==========================================================
        # DESCARGA
        # ==========================================================
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="reporte_gimnasio.xlsx"'

        wb.save(response)
        return response


# ============================================================
# CALENDARIO — TIPO EVENTO, EVENTO CALENDARIO Y ENDPOINT PÚBLICO
# ============================================================

class TipoEventoViewSet(MultiTenantViewSetMixin, viewsets.ModelViewSet):
    queryset = TipoEvento.objects.all()
    serializer_class = TipoEventoSerializer
    permission_classes = [IsAuthenticated, IsAdminUser]


class EventoCalendarioViewSet(MultiTenantViewSetMixin, viewsets.ModelViewSet):
    queryset = EventoCalendario.objects.all()
    serializer_class = EventoCalendarioSerializer
    permission_classes = [IsAuthenticated, IsRecepcionUser]

    def get_queryset(self):
        qs = super().get_queryset().select_related('tipo')
        start = self.request.query_params.get('start')
        end = self.request.query_params.get('end')
        if start and end:
            qs = qs.filter(fecha_inicio__lte=end, fecha_fin__gte=start)
        return qs.order_by('fecha_inicio')

    def perform_create(self, serializer):
        serializer.save(gimnasio=self.request.gimnasio, created_by=self.request.user)


class PublicCalendarioView(APIView):
    """Endpoint público (sin autenticación) con los eventos de un gimnasio."""
    permission_classes = [AllowAny]

    def get(self, request, gimnasio_id):
        gimnasio = get_object_or_404(Gimnasio, pk=gimnasio_id)
        eventos = EventoCalendario.objects.filter(
            gimnasio=gimnasio
        ).select_related('tipo').order_by('fecha_inicio')
        serializer = EventoCalendarioSerializer(eventos, many=True)
        return Response(serializer.data)

# ============================================================
# SOLICITUD DE DEMO
# ============================================================
from .models import DemoRequest
from .serializers import DemoRequestSerializer

class DemoRequestViewSet(viewsets.ModelViewSet):
    """
    Endpoint para recibir solicitudes de demo desde la landing/login.
    POST público (sin autenticar). GET y PATCH solo para admins autenticados.
    """
    queryset = DemoRequest.objects.all()
    serializer_class = DemoRequestSerializer
    http_method_names = ['get', 'post', 'patch', 'options']

    def get_permissions(self):
        if self.request.method == 'POST':
            return [AllowAny()]
        return [IsAuthenticated()]

    def perform_create(self, serializer):
        # Acá a futuro podés agregar lógica para mandarte un email automático
        serializer.save()


# ============================================================
# PLATFORM — SUPERADMIN VIEWS
# ============================================================
from .models import DemoRequest
from .serializers import PlatformStatsSerializer, GimnasioPlatformSerializer, GimnasioPlatformDetailSerializer
from django.db.models import Count, Q, Sum
from datetime import date
from decimal import Decimal


class PlatformStatsView(APIView):
    """Estadísticas globales de la plataforma (solo superadmin)."""
    permission_classes = [IsAuthenticated, IsSuperAdmin]

    def get(self, request):
        today = date.today()
        mes_actual = today.month
        anio_actual = today.year

        # Mes anterior
        if mes_actual == 1:
            prev_month = 12
            prev_year = anio_actual - 1
        else:
            prev_month = mes_actual - 1
            prev_year = anio_actual

        # Total gyms
        total_gimnasios = Gimnasio.objects.count()
        gimnasios_activos = Gimnasio.objects.filter(is_active=True).count()

        # Staff total (admin + recepcion + superadmin)
        total_usuarios_staff = Usuario.objects.count()

        # Demo requests
        demo_pendientes = DemoRequest.objects.filter(estado='pendiente').count()
        demo_contactados = DemoRequest.objects.filter(estado='contactado').count()

        # Ingresos mes global = pagos + diarios (R5)
        pagos_mes = PagoMembresia.objects.filter(
            fecha_pago__month=mes_actual,
            fecha_pago__year=anio_actual
        ).aggregate(total=Sum('monto'))['total'] or Decimal('0')

        diarios_mes = UsuarioGymDay.objects.filter(
            dateInitial__month=mes_actual,
            dateInitial__year=anio_actual
        ).aggregate(total=Sum('price'))['total'] or Decimal('0')

        ingresos_mes_global = pagos_mes + diarios_mes

        # Miembros activos globales (hoy)
        miembros_activos_global = MembresiaAsignada.objects.filter(
            dateInitial__lte=today,
            dateFinal__gte=today
        ).values('miembro').distinct().count()

        # Retención ponderada: SUM(activos_hoy) / SUM(activos_mes_anterior) * 100
        # Miembros activos mes anterior
        if prev_month == 12:
            prev_month_end = date(prev_year + 1, 1, 1) - timedelta(days=1)
        else:
            prev_month_end = date(prev_year, prev_month + 1, 1) - timedelta(days=1)
        prev_month_start = date(prev_year, prev_month, 1)

        activos_hoy = MembresiaAsignada.objects.filter(
            dateInitial__lte=today,
            dateFinal__gte=today
        ).values('miembro', 'miembro__gimnasio').distinct()

        activos_mes_anterior = MembresiaAsignada.objects.filter(
            dateInitial__lte=prev_month_end,
            dateFinal__gte=prev_month_start
        ).values('miembro', 'miembro__gimnasio').distinct()

        # Agrupar por gimnasio para ponderar
        from collections import defaultdict
        activos_hoy_por_gym = defaultdict(int)
        activos_anterior_por_gym = defaultdict(int)

        for a in activos_hoy:
            activos_hoy_por_gym[a['miembro__gimnasio']] += 1
        for a in activos_mes_anterior:
            activos_anterior_por_gym[a['miembro__gimnasio']] += 1

        total_hoy = sum(activos_hoy_por_gym.values())
        total_anterior = sum(activos_anterior_por_gym.values())

        if total_anterior > 0:
            retencion_promedio = Decimal(str(round((total_hoy / total_anterior) * 100, 1)))
        else:
            retencion_promedio = Decimal('100.0')

        data = {
            'total_gimnasios': total_gimnasios,
            'gimnasios_activos': gimnasios_activos,
            'total_usuarios_staff': total_usuarios_staff,
            'demo_pendientes': demo_pendientes,
            'demo_contactados': demo_contactados,
            'ingresos_mes_global': ingresos_mes_global,
            'miembros_activos_global': miembros_activos_global,
            'retencion_promedio': retencion_promedio,
        }
        serializer = PlatformStatsSerializer(data)
        return Response(serializer.data)


class GimnasioPlatformViewSet(viewsets.ModelViewSet):
    """CRUD de gimnasios para superadmin (sin filtro multi-tenant)."""
    permission_classes = [IsAuthenticated, IsSuperAdmin]
    pagination_class = PlatformPagination
    queryset = Gimnasio.objects.all().order_by('-created_at')
    filter_backends = [DjangoFilterBackend, SearchFilter]
    search_fields = ['name', 'address', 'phone']

    def get_serializer_class(self):
        if self.action == 'retrieve':
            return GimnasioPlatformDetailSerializer
        return GimnasioPlatformSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        
        # Anotaciones para list
        if self.action == 'list':
            today = date.today()
            mes_actual = today.month
            anio_actual = today.year

            qs = qs.annotate(
                usuarios_count=Count(
                    'usuarios',
                    filter=Q(usuarios__is_active=True)
                ),
                miembros_activos_count=Count(
                    'miembros__miembro',
                    filter=Q(
                        miembros__miembro__dateInitial__lte=today,
                        miembros__miembro__dateFinal__gte=today
                    ),
                    distinct=True
                ),
                ingresos_mes=(
                    Sum(
                        'miembros__miembro__pagos__monto',
                        filter=Q(
                            miembros__miembro__pagos__fecha_pago__month=mes_actual,
                            miembros__miembro__pagos__fecha_pago__year=anio_actual
                        )
                    ) or Decimal('0')
                ) + (
                    Sum(
                        'miembros_diarios__price',
                        filter=Q(
                            miembros_diarios__dateInitial__month=mes_actual,
                            miembros_diarios__dateInitial__year=anio_actual
                        )
                    ) or Decimal('0')
                )
            )
        return qs